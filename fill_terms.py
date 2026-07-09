# -*- coding: utf-8 -*-
import sys, csv, os
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

DIR=os.path.dirname(os.path.abspath(__file__))
def norm(c):
    s=str(c).strip().split('.')[0].split()[0]
    return str(int(s)) if s.isdigit() else s.upper()

# 建议执行价/敲入价
adv={}
with open(os.path.join(DIR,"建议结构.csv"),encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        try: adv[norm(r["代码"])]=(int(r["建议执行价%"]),int(r["建议敲入价%"]))
        except: pass

FILES=["低波精选组.xlsx","高波精选组.xlsx","市场热度榜.xlsx"]
# 名称映射：从低波+高波借
name_map={}
for fn in ["低波精选组.xlsx","高波精选组.xlsx"]:
    ws=openpyxl.load_workbook(fn,data_only=True).active
    for r in ws.iter_rows(min_row=2,values_only=True):
        if r[1] and r[2]: name_map[norm(r[1])]=r[2]
# 5 个缺失名手动补
name_map.update({"APP":"AppLovin","HOOD":"Robinhood","RDDT":"Reddit",
                 "NET":"Cloudflare","1024":"快手"})

for fn in FILES:
    wb=openpyxl.load_workbook(fn); ws=wb.active
    miss=[]
    for i in range(2,ws.max_row+1):
        region=str(ws.cell(i,1).value or "").strip()
        raw=ws.cell(i,2).value
        if raw is None: continue
        nc=norm(raw)
        # 1. 执行价/敲入价(小数)
        if nc in adv:
            s,k=adv[nc]; ws.cell(i,5).value=round(s/100,2); ws.cell(i,6).value=round(k/100,2)
        else: miss.append(str(raw))
        # 2. 名称(仅市场热度榜且空)
        if fn=="市场热度榜.xlsx" and not ws.cell(i,3).value:
            ws.cell(i,3).value=name_map.get(nc, str(raw))
        # 3. 代码格式 CODE MARKET
        is_hk=region.startswith("港")
        disp=(str(int(nc)) if nc.isdigit() else nc)
        ws.cell(i,2).value=f"{disp} {'HK' if is_hk else 'US'}"
    try:
        wb.save(fn); print(f"✅ {fn}: 填充完成"+(f"; 未匹配执行价: {miss}" if miss else ""))
    except PermissionError:
        print(f"❌ {fn} 被占用(Excel开着), 未保存")
